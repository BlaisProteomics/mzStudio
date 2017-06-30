import xlrd
import openpyxl
import os

def get_sheet_names_xls(filename):
    wb = xlrd.open_workbook(filename)
    sheets = wb.sheets()
    sheetnames = []
    for sheet in sheets:
        sheetnames.append(sheet.name)
    del sheets
    del wb
    return sheetnames

def get_sheet_names_xlsx(filename):
    wb = openpyxl.load_workbook(filename)
    sheet_names = wb.get_sheet_names()
    del wb
    return sheet_names

def get_column_names_xls(filename, sheet_name):
    wb = xlrd.open_workbook(filename)
    sheet = wb.sheet_by_name(sheet_name)
    cols = []
    for col in range(sheet.ncols):
        cols.append(str(sheet.cell(0,col).value))
    del sheet
    del wb
    return cols
    
def get_column_names_xlsx(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheet_name)
    #colnum= sheet.get_highest_column()
    #cols = []
    #for col in range(1, colnum+1):
        #cols.append(str(sheet.cell(row=1,column=col).value))
    cols = [x.value if x.value != None else '' for x in sheet.iter_rows().next() if x]
    del sheet
    del wb
    return cols

def get_sheet_names(filename):
    x, fileExtension = os.path.splitext(filename)
    if fileExtension == '.xls':
        sheet_names = get_sheet_names_xls(filename)
    elif fileExtension == '.xlsx':
        sheet_names = get_sheet_names_xlsx(filename)
    return sheet_names

def get_column_names(filename, sheet_name):
    x, fileExtension = os.path.splitext(filename)
    if fileExtension == '.xls':
        column_names = get_column_names_xls(filename, sheet_name)
    elif fileExtension == '.xlsx':
        column_names = get_column_names_xlsx(filename, sheet_name)
    return column_names